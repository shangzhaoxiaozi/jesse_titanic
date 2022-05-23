import requests
import pandas as pd
import os, random, sys, requests,json,io
from requests_toolbelt.multipart.encoder import MultipartEncoder 
import datetime,time
import pymysql
from apscheduler.schedulers.blocking import BlockingScheduler
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
import warnings
from tempfile import NamedTemporaryFile
warnings.filterwarnings('ignore')


def send_all(fei):

sql = '*****************'

# 读取数据
    db = pymysql.connect(host="********************", user="****", password="****", database="****")
# 使用cursor()方法获取操作游标 
    cursor = db.cursor()
    # SQL 查询语句
    cursor.execute(sql)
   # 获取所有记录列表
    results = cursor.fetchall()
    col = cursor.description
    # print(results)
    cursor.close()
    db.close()


    columns_name = [c[0] for c in col]
    df = pd.DataFrame(results,columns=columns_name)

    df['页面PV'].fillna(0,inplace=True)
    df['UV'].fillna(0,inplace=True)
    df['收藏数'].fillna(0,inplace=True)
    df['加购数'].fillna(0,inplace=True)
#     df['页面PV'] = df['页面PV'].astype('int')
    df['UV'] = df['UV'].astype('int')
    df['收藏数'] = df['收藏数'].astype('int')
    df['加购数'] = df['加购数'].astype('int')

    df['收藏率'] = df['收藏数'] / df['UV']
    df['收藏成本'] = df['总花费'] / df['收藏数']
    df['加购率'] = df['加购数'] / df['UV']
    df['加购成本'] = df['总花费'] / df['加购数']
    df['进店转化率'] = df['成交件数'] / df['UV']
    df['平均单价'] = df['付款金额'] / df['成交件数']
    df['roi'] = df['付款金额'] / df['总花费']

    # filt
    df = df[['date', '序号', '系列', '达人', '抖音ID', '链接素材信息', '视频ID', '推广位', '链接下单信息',
         '卡片出现时间','单价','返点',
       '定向','总成本', '总花费', '代理', '整体播放量',
       '展示数', '点击数', 'CTR', 'CPC', 'CPM', '转化数', '转化成本', '转化率', '有效播放数',
       '有效播放率', '有效播放成本', '进度播放数_25', '进度播放数_50', '进度播放数_75', '进度播放数_99',
       '进度播放率_25', '进度播放率_50', '进度播放率_75', '进度播放率_99', '点赞数', '评论数', '分享数',
       '成交件数', '付款金额',  'UV', '收藏数', '加购数', '收藏率', '收藏成本', '加购率', '加购成本', '进店转化率',
       '平均单价', 'roi']]

    # sort_order
    df= df[['date', '系列', '达人', '抖音ID', '链接素材信息', '视频ID','代理', '推广位', '链接下单信息',
       '定向', '总成本', '整体播放量', '展示数', '点击数', 'CTR', '总花费',  'CPC', 'CPM',
       '转化数', '转化成本', '转化率',
 'UV', '收藏数', '加购数','收藏率', '收藏成本', '加购率',
 '加购成本', '成交件数','进店转化率', '付款金额',  '平均单价', 'roi',
'有效播放数', '有效播放率', '有效播放成本', '进度播放数_25',
       '进度播放数_50', '进度播放数_75', '进度播放数_99', '进度播放率_25', '进度播放率_50', '进度播放率_75',
       '进度播放率_99', '点赞数', '评论数', '分享数']]


# int 
    int_list = ['总成本', '整体播放量', '展示数', '点击数', '总花费', '转化数',
 'UV', '收藏数', '加购数', '成交件数',
       '付款金额',  '有效播放数',   '进度播放数_25',
       '进度播放数_50', '进度播放数_75', '进度播放数_99',
        '点赞数', '评论数', '分享数']
    for i in int_list:
        df[i].fillna(0,inplace=True)
        df[i] = df[i].astype('int')
    
    
# float-2
    float_list = ['CPC', 'CPM', '转化成本',  '收藏成本', '加购成本','平均单价','roi','有效播放成本']
    for i in float_list:
        df[i].fillna(0,inplace=True)
        df[i] = df[i].round(2)
    
# %
    percent_list = ['CTR', '转化率',  '收藏率',  '加购率', 
       '进店转化率', '有效播放率', 
        '进度播放率_25', '进度播放率_50', '进度播放率_75',
       '进度播放率_99']
    for i in percent_list:
        df[i].fillna(0,inplace=True)
        df[i] = df[i].apply(lambda x:format(x,'.2%'))
    # add %
    df.columns = ['date', '系列', '达人', '抖音ID', '链接素材信息', '视频ID', '代理', '推广位', '链接下单信息',
       '定向', '总成本', '整体播放量', '展示数', '点击数', 'CTR', '总花费', 'CPC', 'CPM', '转化数',
       '转化成本', '转化率', 'UV', '收藏数', '加购数', '收藏率', '收藏成本', '加购率', '加购成本', '成交件数',
       '进店转化率', '付款金额', '平均单价', 'roi', '有效播放数', '有效播放率', '有效播放成本', '进度播放数_25%',
       '进度播放数_50%', '进度播放数_75%', '进度播放数_99%', '进度播放率_25%', '进度播放率_50%', '进度播放率_75%',
       '进度播放率_99%', '点赞数', '评论数', '分享数']


    wb1 = Workbook()

    df['date_m'] = df['date'].apply(lambda x:str(x)[:7])
    sheet_list =df['date_m'].unique().tolist()

    for i in sheet_list:
        wb1.create_sheet(i)
    del wb1['Sheet']

    # 分表写入
    for sheet in sheet_list:
        ws = wb1[sheet]
        df_tmp = df[df['date_m'] ==  sheet]
        del df_tmp['date_m']
        df_tmp.reset_index(drop=True,inplace=True)
    
        # start_position
        
            # start_position
        start_row = 3
        start_col = 'B'
        insert_data_mul(kw={'row':2,'col':'C','value':'素材信息','length':5,'color':'FFEBCD','worksheet':ws})
        insert_data_mul(kw={'row':2,'col':'H','value':'下单信息','length':5,'color':'6495ED','worksheet':ws})
        insert_data_mul(kw={'row':2,'col':'M','value':'前端展示数据','length':7,'color':'8FBC8F','worksheet':ws})
        insert_data_mul(kw={'row':2,'col':'T','value':'落地页转化数据','length':3,'color':'E9967A','worksheet':ws})
        insert_data_mul(kw={'row':2,'col':'W','value':'即时电商数据','length':12,'color':'1E90FF','worksheet':ws})
        insert_data_mul(kw={'row':2,'col':'AI','value':'素材数据','length':14,'color':'ADFF2F','worksheet':ws})

        data_insert(start_row,start_col,df_tmp,ws)
        juzhong(ws)

        # hebing date from vertical
        b = 4
    #   e = -1
        for i in range(5,ws.max_row):
            if ws['B'+str(b)].value != ws['B'+str(i)].value:
                e = i-1
                # hebing tmp
                ws.merge_cells('B'+str(b)+':'+'B'+str(e))
                b=i
    # last set of date
        ws.merge_cells('B'+str(b)+':'+'B'+str(ws.max_row))     

    # the last work 
        ws['B2'] = '下单日期'
        ws['B3'] = '下单日期'
        ws.merge_cells('B2:B3')

        ws.column_dimensions['B'].width=16
        ws.column_dimensions['D'].width=16
        ws.column_dimensions['F'].width=16
        ws.column_dimensions['J'].width=16

        dimen_list = ['AI','AJ',
 'AK',
 'AL',
 'AM',
 'AN',
 'AO',
 'AP',
 'AQ',
 'AR',
 'AS']
        for i in dimen_list:
            ws.column_dimensions[i].width=16
    # output
    date_last = df.iloc[df['date'].index.size-1,0]
    date_last = date_last.isoformat()
    file_name = 'MOODY抖音内容热推投放日报_'+date_last+'.xlsx'
#     wb1.save(file_name)
    print('数据格式处理完成')
    
    with NamedTemporaryFile() as tmp:
        wb1.save(tmp.name)
        output = io.BytesIO(tmp.read())
   

    print('***********************************************************')

    #先发送消息提醒***********************************************************
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
    req_body = {
        "receive_id": fei.get_chat_list()['chat_id'],
        "content": " {\"text\":\"<at user_id=\\\"all\\\"></at> MOODY抖音内容热推投放日报-更新提醒\"}",
        "msg_type": "text"
    }
    res=requests.post("https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id", headers=fei.headers,json=req_body)
    print('MOODY抖音内容：消息已发送')

    #发送文件***********************************************************
    url = 'https://open.feishu.cn/open-apis/im/v1/files' 
    headers = fei.headers


    multipart_encoder = MultipartEncoder(
    fields = #这里根据服务器需要的参数格式进行修改
            {
            'file_type':'stream',
            'file_name':file_name,
            'duration':'3000',
            'file': ('file', output, 'application/octet-stream')
            },
        boundary='---7MA4YWxkTrZu0gW'
    )


    headers['Content-Type'] = multipart_encoder.content_type
        #请求头必须包含一个特殊的头信息,类似于Content-Type: multipart/form-data; boundary=${bound}
        #注意：这里请求头也可以自己设置Content-Type信息，用于自定义boundary
    r = requests.post(url, data=multipart_encoder, headers=headers)

    file_key = r.content.decode()
    file_key = json.loads(file_key)

    req_body = {
        "receive_id": fei.get_chat_list()['chat_id'],
        "content": json.dumps(file_key['data']),
        "msg_type": "file"
    }

    res=requests.post("https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id", headers=fei.headers,json=req_body)
#     print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
    print('MOODY抖音内容热推：文件发送成功！')    


def ini_all(chat_name):
    app_id="****************************"
    app_secret = '****************************'
    APP_VERIFICATION_TOKEN = "****************************"

    ##########################################
    fei=FeishuApi(app_id,app_secret,chat_name)
    send_all(fei)