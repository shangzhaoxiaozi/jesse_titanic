import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
import warnings
warnings.filterwarnings('ignore')




# excel columns name make,from A  to ZZ
def excel_columns_make():
    chr_list = [chr(i) for i in range(65, 91)]
    chr_list_all = []
    for i in chr_list:
        for j in chr_list:
            chr_list_all.append(i + j)
    chr_list_all = chr_list + chr_list_all
    return chr_list_all


# return assigned chr_list,from chr_,length of num
# the returned chr list is without row number,such as : [A,B,....]
def return_asigned_chr_list(chr_,num,chr_list_all):
    for i in range(len(chr_list_all)):
        if chr_list_all[i] == chr_:
            return chr_list_all[i:i + num]


# dataframe insert cell one by one,columns first,contents then
# the df_tmp passed should by reset_index,like (0,1,2,3,4,5.....)
def dataframe_data_insert(**kw):
    if 'start_row' not in kw.keys():
        start_row = 1
    else:
        start_row = kw['start_row']

    if 'start_col' not in kw.keys():
        start_col = 'A'
    else:
        start_col = kw['start_col']

    ws = kw['worksheet']
    df_tmp = kw['dataframe']
    chr_list_all = kw['chr_all']

    # reset_index again
    df_tmp.reset_index(drop=True, inplace=True)
    columns_name = df_tmp.columns.values
    num = len(columns_name)
    columns_name_index = return_asigned_chr_list(start_col,num,chr_list_all)
    # columns name insert
    columns_p = [i + str(start_row) for i in columns_name_index]
    for i in range(len(columns_p)):
        ws[columns_p[i]] = columns_name[i]
    start_row = start_row + 1

    # contents insert
    for row_i in df_tmp.index:
        row_i_position = [i + str(start_row) for i in columns_name_index]
        for i in range(len(row_i_position)):
            ws[row_i_position[i]] = df_tmp.iloc[row_i, i]
        start_row = start_row + 1
    return ws

# whole worksheet style set,alignment,font
def whole_ws_style_set(**kw):
    ws = kw['worksheet']
    nrows = ws.max_row  # 获得行数
    ncols = ws.max_column
    for i in range(nrows):
        for j in range(ncols):
            ws.cell(row=i + 1, column=j + 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=i + 1, column=j + 1).font = Font(name=u'微软雅黑')
    return ws

# free_style_data_insert(row=2,col='C',value='素材信息',length=5,color='FFEBCD')
# free_style_data_insert(row=2,col='C',value='素材信息',length=5)
# ws = openpyxl_format.free_style_data_insert(row=1,col='B',value='moody旗舰店',length=4,color='FFEBCD',worksheet=ws,chr_all=chr_list_all)
# ws = openpyxl_format.free_style_data_insert(row=1,col='F',value='moody品牌店',length=4,color='6495ED',worksheet=ws,chr_all=chr_list_all)
# ws = openpyxl_format.free_style_data_insert(row=1,col='J',value='moody彩瞳小店',length=4,color='8FBC8F',worksheet=ws,chr_all=chr_list_all)
# ws = openpyxl_format.free_style_data_insert(row=1,col='M',value='total',length=2,color='E9967A',worksheet=ws,chr_all=chr_list_all)

def free_style_data_insert(**kw):
    ws = kw['worksheet']
    chr_list_all = kw['chr_all']
    row = kw['row']
    list_t = return_asigned_chr_list(kw['col'], kw['length'],chr_list_all)
    for i in list_t:
        ws[i + str(kw['row'])] = kw['value']
    # hebing_horizon by the way
    b = list_t[0] + str(kw['row'])
    e = list_t[-1] + str(kw['row'])
    ws.merge_cells(b + ':' + e)
    # if kw contain color ,set back color by the way
    if 'color' in kw.keys():
        color = kw['color']
        for j in list_t:
            fill = PatternFill("solid", fgColor=color)
            ws[j + str(row)].fill = fill
            ws[j + str(row + 1)].fill = fill

    return ws